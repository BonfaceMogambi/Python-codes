import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference

def create_weekly_report(data_file, output_name="Weekly_Report"):
    """
    Generate a formatted Excel report with charts
    """
    
    # Read your data (works with CSV, Excel, etc.)
    df = pd.read_csv(data_file)
    
    # Create a new Excel file
    timestamp = datetime.now().strftime("%Y-%m-%d")
    output_file = f"{output_name}_{timestamp}.xlsx"
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write summary statistics
        summary = df.describe()
        summary.to_excel(writer, sheet_name='Summary')
        
        # Write full data
        df.to_excel(writer, sheet_name='Data', index=False)
        
        # Access the workbook to add formatting
        workbook = writer.book
        summary_sheet = workbook['Summary']
        data_sheet = workbook['Data']
        
        # Format headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in data_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in data_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            data_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add a chart if we have numeric data
        if len(df.select_dtypes(include=['number']).columns) > 0:
            chart_sheet = workbook.create_sheet('Charts')
            chart = BarChart()
            chart.title = "Data Overview"
            chart.style = 10
            
            # Use first 10 rows for the chart
            data = Reference(data_sheet, min_col=2, min_row=1, max_row=min(11, len(df)+1), max_col=len(df.columns))
            categories = Reference(data_sheet, min_col=1, min_row=2, max_row=min(11, len(df)+1))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart_sheet.add_chart(chart, "A1")
    
    print(f"✓ Report created: {output_file}")
    return output_file
# HOW TO USE:
# Put your data in a CSV file, then run:
# create_weekly_report("sales_data.csv", "Sales_Report")